from __future__ import annotations

import html
import io
import re
import unicodedata
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st


APP_DIR = Path(__file__).resolve().parent
DEFAULT_FILE = APP_DIR / "data" / "Analise CD 989.xlsx"
CATEGORIES = ["Ajustes Lojas", "Transferências", "Devoluções", "Inventário"]


def normalize_label(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = re.sub(r"\s+", " ", text).strip().casefold()
    return text


def normalize_code(value: object) -> str | None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value).strip()
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\.0$", "", text)


def first_valid(series: pd.Series) -> str:
    values = series.dropna().astype(str).str.strip()
    values = values[values.ne("")]
    return values.iloc[0] if not values.empty else "Produto sem descrição"


def read_contiguous_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> pd.DataFrame:
    """Lê apenas a área útil e evita percorrer linhas formatadas até 1.048.576."""
    rows: list[tuple] = []
    empty_streak = 0
    for row in ws.iter_rows(values_only=True):
        is_empty = all(value is None for value in row)
        if is_empty:
            empty_streak += 1
        else:
            empty_streak = 0
            rows.append(row)
        if rows and empty_streak >= 25:
            break

    if not rows:
        return pd.DataFrame()

    headers = [str(value).strip() if value is not None else f"coluna_{index}" for index, value in enumerate(rows[0])]
    frame = pd.DataFrame(rows[1:], columns=headers).dropna(how="all")
    frame.columns = [str(column).strip() for column in frame.columns]
    return frame


@st.cache_data(show_spinner=False)
def load_workbook(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    frames: dict[str, pd.DataFrame] = {}
    try:
        for sheet_name in workbook.sheetnames:
            frames[sheet_name.strip()] = read_contiguous_sheet(workbook[sheet_name])
    finally:
        workbook.close()
    return frames


def find_sheet(frames: dict[str, pd.DataFrame], expected: str) -> pd.DataFrame:
    normalized_expected = normalize_label(expected)
    for name, frame in frames.items():
        if normalize_label(name) == normalized_expected:
            return frame.copy()
    raise ValueError(f'A aba obrigatória "{expected}" não foi encontrada.')


def standardize_columns(frame: pd.DataFrame) -> pd.DataFrame:
    aliases = {
        "codigo produto": "codigo_produto",
        "produto": "produto",
        "quantidade em estoque": "quantidade_estoque",
        "qtd diferenca": "quantidade_mov",
        "valor custo liquido": "valor_liquido",
        "valor custo bruto": "valor_bruto",
        "empresa": "empresa",
        "c.g.o.": "cgo",
        "codigo c.g.o.": "codigo_cgo",
        "codigo empresa": "codigo_empresa",
    }
    renamed = {}
    for column in frame.columns:
        normalized = normalize_label(column)
        renamed[column] = aliases.get(normalized, normalized.replace(" ", "_"))
    return frame.rename(columns=renamed)


def numeric(series: pd.Series | None, index: pd.Index) -> pd.Series:
    if series is None:
        return pd.Series(float("nan"), index=index, dtype="float64")
    return pd.to_numeric(series, errors="coerce")


@st.cache_data(show_spinner=False)
def prepare_data(file_bytes: bytes) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    frames = load_workbook(file_bytes)
    stock_raw = standardize_columns(find_sheet(frames, "Estoque CD"))
    required_stock = {"codigo_produto", "produto", "quantidade_estoque", "valor_liquido"}
    missing_stock = required_stock.difference(stock_raw.columns)
    if missing_stock:
        raise ValueError("Colunas ausentes em Estoque CD: " + ", ".join(sorted(missing_stock)))

    stock_raw["codigo_produto"] = stock_raw["codigo_produto"].map(normalize_code)
    stock_raw["quantidade_estoque"] = numeric(stock_raw["quantidade_estoque"], stock_raw.index).fillna(0).abs()
    stock_raw["valor_quebra"] = numeric(stock_raw["valor_liquido"], stock_raw.index).fillna(0).abs()
    stock_raw = stock_raw.dropna(subset=["codigo_produto"])

    stock = (
        stock_raw.groupby("codigo_produto", as_index=False)
        .agg(
            produto=("produto", first_valid),
            quantidade_estoque=("quantidade_estoque", "sum"),
            valor_quebra=("valor_quebra", "sum"),
        )
    )
    cd_codes = set(stock["codigo_produto"])

    specs = {
        "Ajustes Lojas": ("Ajustes Lojas", "empresa"),
        "Transferências": ("Transferências", "cgo"),
        "Devoluções": ("Devoluções", "cgo"),
        "Inventário": ("Inventário", "cgo"),
    }
    movement_frames: list[pd.DataFrame] = []
    quality_rows: list[dict] = []

    quality_rows.append(
        {
            "Aba": "Estoque CD",
            "Linhas úteis": len(stock_raw),
            "Produtos únicos": stock_raw["codigo_produto"].nunique(),
            "Produtos do CD vinculados": stock_raw["codigo_produto"].nunique(),
            "Produtos ignorados": 0,
        }
    )

    for category, (sheet_name, subcategory_column) in specs.items():
        raw = standardize_columns(find_sheet(frames, sheet_name))
        required = {"codigo_produto", "produto", "valor_liquido"}
        missing = required.difference(raw.columns)
        if missing:
            raise ValueError(f"Colunas ausentes em {sheet_name}: " + ", ".join(sorted(missing)))

        raw["codigo_produto"] = raw["codigo_produto"].map(normalize_code)
        all_codes = set(raw["codigo_produto"].dropna())
        linked = raw[raw["codigo_produto"].isin(cd_codes)].copy()
        linked["categoria"] = category
        linked["valor_movimento_assinado"] = numeric(linked.get("valor_liquido"), linked.index).fillna(0)
        linked["valor_movimento"] = linked["valor_movimento_assinado"].abs()
        linked["quantidade_movimento_assinada"] = numeric(linked.get("quantidade_mov"), linked.index)
        linked["quantidade_movimento"] = linked["quantidade_movimento_assinada"].abs()

        if subcategory_column in linked.columns:
            linked["subcategoria"] = linked[subcategory_column].fillna("Não informado").astype(str).str.strip()
            linked.loc[linked["subcategoria"].eq(""), "subcategoria"] = "Não informado"
        else:
            linked["subcategoria"] = "Sem detalhamento"

        movement_frames.append(
            linked[
                [
                    "codigo_produto",
                    "produto",
                    "categoria",
                    "subcategoria",
                    "quantidade_movimento",
                    "quantidade_movimento_assinada",
                    "valor_movimento",
                    "valor_movimento_assinado",
                ]
            ]
        )
        quality_rows.append(
            {
                "Aba": category,
                "Linhas úteis": len(raw),
                "Produtos únicos": len(all_codes),
                "Produtos do CD vinculados": linked["codigo_produto"].nunique(),
                "Produtos ignorados": len(all_codes.difference(cd_codes)),
            }
        )

    movements = pd.concat(movement_frames, ignore_index=True) if movement_frames else pd.DataFrame()
    quality = pd.DataFrame(quality_rows)
    return stock, movements, quality


def brl(value: float) -> str:
    formatted = f"{float(value):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {formatted}"


def integer_br(value: float) -> str:
    return f"{float(value):,.0f}".replace(",", ".")


def percentage(value: float) -> str:
    return f"{value * 100:.1f}%".replace(".", ",")


def metric_card(label: str, value: str, detail: str = "") -> None:
    st.markdown(
        f"""
        <div class="metric-card">
            <div class="metric-label">{html.escape(label)}</div>
            <div class="metric-value">{html.escape(value)}</div>
            <div class="metric-detail">{html.escape(detail)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def build_ranking(stock: pd.DataFrame, movements: pd.DataFrame) -> pd.DataFrame:
    if movements.empty:
        movement_total = pd.Series(dtype="float64", name="movimentacao_vinculada")
    else:
        movement_total = (
            movements.groupby("codigo_produto")["valor_movimento_assinado"]
            .sum()
            .rename("movimentacao_vinculada")
        )

    ranking = stock.merge(movement_total, on="codigo_produto", how="left")
    ranking["movimentacao_vinculada"] = ranking["movimentacao_vinculada"].fillna(0)
    ranking["valor_identificado"] = ranking["movimentacao_vinculada"]
    ranking["saldo_justificar"] = (ranking["valor_quebra"] - ranking["valor_identificado"]).clip(lower=0)
    ranking["cobertura"] = ranking["valor_identificado"].div(ranking["valor_quebra"].replace(0, pd.NA)).fillna(0)
    ranking["status"] = "Identificada"
    ranking.loc[ranking["movimentacao_vinculada"].eq(0), "status"] = "Sem Ajustes"
    ranking.loc[ranking["movimentacao_vinculada"].lt(0), "status"] = "Ajuste negativo"
    ranking.loc[
        ranking["movimentacao_vinculada"].gt(0) & ranking["movimentacao_vinculada"].lt(ranking["valor_quebra"]),
        "status",
    ] = "Parcial"
    return ranking


def inject_css() -> None:
    st.markdown(
        """
        <style>
        :root {
            color-scheme: dark;
            --accent: #ff4b55;
            --app-bg: #0d1117;
            --sidebar-bg: #232630;
            --panel: #191d24;
            --panel-secondary: #14181e;
            --input-bg: #11151b;
            --text: #f3f4f6;
            --text-strong: #f8fafc;
            --muted: #aeb5c0;
            --muted-soft: #8f98a5;
            --border: #2c323d;
            --button-border: #3a414d;
            --shadow: rgba(0, 0, 0, .18);
            --hover: rgba(255, 75, 85, .055);
        }
        @media (prefers-color-scheme: light) {
            :root {
                color-scheme: light;
                --app-bg: #f6f7fb;
                --sidebar-bg: #eef0f5;
                --panel: #ffffff;
                --panel-secondary: #f8f9fc;
                --input-bg: #ffffff;
                --text: #1b1f27;
                --text-strong: #111827;
                --muted: #5f6877;
                --muted-soft: #70798a;
                --border: #d8dde6;
                --button-border: #c5cbd6;
                --shadow: rgba(25, 35, 55, .08);
                --hover: rgba(255, 75, 85, .07);
            }
        }
        html, body, [data-testid="stAppViewContainer"], .stApp {
            background: var(--app-bg) !important;
            color: var(--text) !important;
        }
        [data-testid="stHeader"] {
            background: color-mix(in srgb, var(--app-bg) 92%, transparent) !important;
            color: var(--text) !important;
        }
        [data-testid="stToolbar"], [data-testid="stDecoration"] {
            color: var(--text) !important;
        }
        [data-testid="stSidebar"] {
            background: var(--sidebar-bg) !important;
            color: var(--text) !important;
            border-right: 1px solid var(--border);
        }
        [data-testid="stSidebar"] > div { padding-top: 1.1rem; }
        .stApp p, .stApp label, .stApp h1, .stApp h2, .stApp h3,
        .stApp h4, .stApp h5, .stApp h6, [data-testid="stSidebar"] p,
        [data-testid="stSidebar"] label { color: var(--text); }
        .block-container {
            max-width: none !important;
            padding: 1.35rem 2rem 3rem !important;
        }
        h1, h2, h3 { letter-spacing: -0.025em; }
        .eyebrow {
            display: block; min-height: 1.8rem; padding: .25rem 0 .45rem;
            overflow: visible; line-height: 1.4; color: var(--accent);
            font-weight: 700; text-transform: uppercase; letter-spacing: .12em; font-size: .72rem;
        }
        .hero-subtitle { color: var(--muted); margin-top: -.55rem; margin-bottom: 1.35rem; }
        .metric-card {
            min-width: 0; min-height: 132px; padding: 18px 18px 16px; overflow: hidden;
            border: 1px solid var(--border);
            border-radius: 12px; background: linear-gradient(145deg, var(--panel) 0%, var(--panel-secondary) 100%);
            box-shadow: 0 10px 26px var(--shadow);
        }
        .metric-label { color: var(--muted); font-size: .78rem; font-weight: 650; min-height: 20px; }
        .metric-value {
            color: var(--text-strong); font-size: clamp(.92rem, 1.28vw, 1.45rem); line-height: 1.15;
            letter-spacing: -.035em; font-weight: 720; font-variant-numeric: tabular-nums;
            margin: 9px 0 6px; white-space: nowrap;
        }
        .metric-detail { color: var(--muted-soft); font-size: .72rem; }
        [data-testid="stDataFrame"] {
            border: 1px solid var(--border); border-radius: 10px; overflow: hidden;
            background: var(--panel);
        }
        [data-testid="stExpander"] {
            border: 1px solid var(--border); border-radius: 10px;
            background: linear-gradient(145deg, var(--panel) 0%, var(--panel-secondary) 100%);
            overflow: hidden;
        }
        [data-testid="stExpander"] summary { padding: .72rem .9rem; }
        [data-testid="stExpander"] summary:hover { background: var(--hover); }
        [data-baseweb="tab-list"] { gap: 1.1rem; border-bottom: 1px solid var(--border); }
        [data-baseweb="tab"] { padding-left: 0; padding-right: 0; }
        [data-baseweb="tab"][aria-selected="true"] { color: var(--accent) !important; }
        [data-baseweb="input"] > div, [data-baseweb="select"] > div,
        [data-testid="stFileUploaderDropzone"], textarea, input {
            background: var(--input-bg) !important;
            color: var(--text) !important;
            border-color: var(--border) !important;
        }
        [data-baseweb="popover"], [role="listbox"] {
            background: var(--panel) !important;
            color: var(--text) !important;
        }
        .stButton > button, .stDownloadButton > button {
            color: var(--text); background: var(--panel); border-color: var(--button-border); border-radius: 8px;
        }
        .stButton > button:hover, .stDownloadButton > button:hover { border-color: var(--accent); color: var(--accent); }
        [data-testid="stFileUploaderDropzone"] { background: var(--input-bg); border-color: var(--border); }
        .detail-title { border-left: 3px solid var(--accent); padding-left: 12px; margin: 1.2rem 0 .8rem; }
        .detail-code { color: var(--muted-soft); }
        .method-note { color: var(--muted-soft); font-size: .78rem; line-height: 1.45; }
        hr { border-color: var(--border) !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def detail_table(movements: pd.DataFrame, category: str) -> pd.DataFrame:
    category_data = movements[movements["categoria"].eq(category)]
    if category_data.empty:
        return pd.DataFrame()
    grouped = (
        category_data.groupby("subcategoria", as_index=False)
        .agg(
            quantidade=("quantidade_movimento_assinada", lambda values: values.sum(min_count=1)),
            valor=("valor_movimento_assinado", "sum"),
            registros=("codigo_produto", "size"),
        )
        .sort_values("valor", ascending=False)
    )
    grouped["Quantidade"] = grouped["quantidade"].map(lambda value: "—" if pd.isna(value) else integer_br(value))
    grouped["Valor vinculado"] = grouped["valor"].map(brl)
    return grouped.rename(columns={"subcategoria": "Detalhamento", "registros": "Registros"})[
        ["Detalhamento", "Quantidade", "Valor vinculado", "Registros"]
    ]


def render_store_ranking(
    stock: pd.DataFrame,
    movements: pd.DataFrame,
    search: str,
    top_n: int,
) -> None:
    adjustments = movements[movements["categoria"].eq("Ajustes Lojas")].copy()
    if adjustments.empty:
        st.warning("Nenhum registro foi encontrado na aba Ajustes Lojas para os produtos do Estoque CD.")
        return

    totals = (
        adjustments.groupby("codigo_produto", as_index=False)
        .agg(
            valor_ajustes=("valor_movimento_assinado", "sum"),
            quantidade_ajustada=("quantidade_movimento_assinada", lambda values: values.sum(min_count=1)),
            lojas=("subcategoria", "nunique"),
            registros=("codigo_produto", "size"),
        )
    )
    product_ranking = stock.merge(totals, on="codigo_produto", how="left")
    product_ranking[["valor_ajustes", "lojas", "registros"]] = product_ranking[
        ["valor_ajustes", "lojas", "registros"]
    ].fillna(0)

    if search.strip():
        token = normalize_label(search)
        mask = product_ranking.apply(
            lambda row: token in normalize_label(row["produto"])
            or token in normalize_label(row["codigo_produto"]),
            axis=1,
        )
        product_ranking = product_ranking[mask]

    controls = st.columns([1.35, 1])
    with controls[0]:
        store_order = st.selectbox(
            "Ordenar produtos por",
            ["Maior valor no Estoque CD", "Maior valor em Ajustes Lojas", "Maior quantidade em estoque"],
            key="store_ranking_order",
        )
    with controls[1]:
        show_without_adjustments = st.toggle(
            "Exibir produtos sem ajustes",
            value=False,
            key="show_without_store_adjustments",
        )

    if not show_without_adjustments:
        product_ranking = product_ranking[product_ranking["valor_ajustes"].gt(0)]

    order_columns = {
        "Maior valor no Estoque CD": "valor_quebra",
        "Maior valor em Ajustes Lojas": "valor_ajustes",
        "Maior quantidade em estoque": "quantidade_estoque",
    }
    order_column = order_columns[store_order]
    product_ranking = (
        product_ranking.sort_values([order_column, "valor_ajustes"], ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )

    total_adjustments = adjustments["valor_movimento_assinado"].sum()
    adjusted_products = totals["codigo_produto"].nunique()
    adjusted_stores = adjustments["subcategoria"].nunique()
    summary_columns = st.columns(3)
    with summary_columns[0]:
        metric_card("Total Ajustes Lojas", brl(total_adjustments), "somente produtos do Estoque CD")
    with summary_columns[1]:
        metric_card("Produtos com ajustes", integer_br(adjusted_products), "produtos vinculados")
    with summary_columns[2]:
        metric_card("Lojas identificadas", integer_br(adjusted_stores), "lojas com movimentação")

    st.markdown("#### Produtos e ajustes por loja")
    st.caption("Clique na seta de um produto para abrir o ranking das lojas e os valores ajustados.")

    if product_ranking.empty:
        st.info("Nenhum produto corresponde aos filtros atuais.")
        return

    for position, product in product_ranking.iterrows():
        label = (
            f'**{position + 1}. {product["produto"]}**  ·  Código {product["codigo_produto"]}'
            f'  ·  Estoque CD: **{brl(product["valor_quebra"])}**'
            f'  ·  Ajustes Lojas: **{brl(product["valor_ajustes"])}**'
        )
        with st.expander(label):
            product_stores = adjustments[
                adjustments["codigo_produto"].eq(product["codigo_produto"])
            ]
            if product_stores.empty:
                st.info("Este produto não possui ajustes por loja.")
                continue

            store_table = (
                product_stores.groupby("subcategoria", as_index=False)
                .agg(
                    quantidade_ajustada=("quantidade_movimento_assinada", lambda values: values.sum(min_count=1)),
                    valor_ajustes=("valor_movimento_assinado", "sum"),
                    registros=("codigo_produto", "size"),
                )
                .sort_values("valor_ajustes", ascending=False)
                .reset_index(drop=True)
            )
            store_table.insert(0, "ranking", range(1, len(store_table) + 1))
            store_table["participacao"] = store_table["valor_ajustes"].div(
                store_table["valor_ajustes"].sum()
            )
            store_table = store_table.rename(
                columns={
                    "ranking": "#",
                    "subcategoria": "Loja",
                    "quantidade_ajustada": "Quantidade ajustada",
                    "valor_ajustes": "Valor Ajustes Lojas",
                    "participacao": "% dos ajustes",
                    "registros": "Registros",
                }
            )
            st.dataframe(
                store_table,
                hide_index=True,
                width="stretch",
                height=min(430, 39 + 35 * max(len(store_table), 1)),
                column_config={
                    "#": st.column_config.NumberColumn(width="small", format="%d"),
                    "Loja": st.column_config.TextColumn(width="large"),
                    "Quantidade ajustada": st.column_config.NumberColumn(format="%,.0f"),
                    "Valor Ajustes Lojas": st.column_config.NumberColumn(format="R$ %,.2f"),
                    "% dos ajustes": st.column_config.NumberColumn(format="percent"),
                    "Registros": st.column_config.NumberColumn(format="%d"),
                },
            )


def render_informative_movements(
    stock: pd.DataFrame,
    movements: pd.DataFrame,
    search: str,
    top_n: int,
) -> None:
    informative = movements[~movements["categoria"].eq("Ajustes Lojas")].copy()
    st.markdown("#### Movimentações informativas")
    st.caption(
        "Transferências, devoluções e inventário. "
    )

    cards: list[tuple[str, float, str]] = []
    for category in ["Transferências", "Devoluções"]:
        category_data = informative[informative["categoria"].eq(category)]
        cards.append(
            (
                f"Total {category}",
                float(category_data["valor_movimento_assinado"].sum()),
                f"{category_data['codigo_produto'].nunique()} produtos do CD",
            )
        )

    inventory = informative[informative["categoria"].eq("Inventário")]
    for cgo, cgo_data in inventory.groupby("subcategoria", sort=True):
        cards.append(
            (
                str(cgo).title(),
                float(cgo_data["valor_movimento_assinado"].sum()),
                f"{cgo_data['codigo_produto'].nunique()} produtos do CD",
            )
        )

    if cards:
        card_columns = st.columns(len(cards))
        for column, (label, value, detail) in zip(card_columns, cards):
            with column:
                metric_card(label, brl(value), detail)

    comparison = stock[["codigo_produto", "produto", "valor_quebra"]].copy()
    movement_columns: list[str] = []

    for category in ["Transferências", "Devoluções"]:
        column_name = category
        totals = (
            informative[informative["categoria"].eq(category)]
            .groupby("codigo_produto")["valor_movimento_assinado"]
            .sum()
            .rename(column_name)
        )
        comparison = comparison.merge(totals, on="codigo_produto", how="left")
        movement_columns.append(column_name)

    if not inventory.empty:
        inventory_pivot = inventory.pivot_table(
            index="codigo_produto",
            columns="subcategoria",
            values="valor_movimento_assinado",
            aggfunc="sum",
            fill_value=0,
        )
        renamed_inventory = {}
        for column in inventory_pivot.columns:
            normalized = normalize_label(column)
            if "entrada" in normalized:
                renamed_inventory[column] = "Inventário - Entrada"
            elif "saida" in normalized:
                renamed_inventory[column] = "Inventário - Saída"
            else:
                renamed_inventory[column] = f"Inventário - {column}"
        inventory_pivot = inventory_pivot.rename(columns=renamed_inventory).reset_index()
        comparison = comparison.merge(inventory_pivot, on="codigo_produto", how="left")
        movement_columns.extend(renamed_inventory.values())

    comparison[movement_columns] = comparison[movement_columns].fillna(0)
    comparison["Total informativo"] = comparison[movement_columns].sum(axis=1)

    if search.strip():
        token = normalize_label(search)
        comparison = comparison[
            comparison.apply(
                lambda row: token in normalize_label(row["produto"])
                or token in normalize_label(row["codigo_produto"]),
                axis=1,
            )
        ]

    comparison["_ordem"] = comparison["Total informativo"].abs()
    comparison = comparison.sort_values("_ordem", ascending=False).drop(columns="_ordem").head(top_n)
    comparison.insert(0, "ranking", range(1, len(comparison) + 1))
    comparison = comparison.rename(
        columns={
            "ranking": "#",
            "codigo_produto": "Código",
            "produto": "Produto",
            "valor_quebra": "Valor líquido Estoque CD",
        }
    )

    st.markdown("#### Movimentações por produto")
    column_config = {
        "#": st.column_config.NumberColumn(width="small", format="%d"),
        "Código": st.column_config.TextColumn(width="small"),
        "Produto": st.column_config.TextColumn(width="large"),
        "Valor líquido Estoque CD": st.column_config.NumberColumn(format="R$ %,.2f"),
        "Total informativo": st.column_config.NumberColumn(format="R$ %,.2f"),
    }
    for column in movement_columns:
        column_config[column] = st.column_config.NumberColumn(format="R$ %,.2f")

    st.dataframe(
        comparison,
        hide_index=True,
        width="stretch",
        height=min(720, 39 + 35 * max(len(comparison), 1)),
        column_config=column_config,
    )


def render_product_detail(selected: pd.Series, product_movements: pd.DataFrame) -> None:
    st.markdown(
        f'<div class="detail-title"><h3 style="margin:0">{html.escape(str(selected["produto"]))}</h3>'
        f'<span class="detail-code">Código {html.escape(str(selected["codigo_produto"]))}</span></div>',
        unsafe_allow_html=True,
    )

    columns = st.columns(4)
    with columns[0]:
        metric_card("Quantidade no CD", integer_br(selected["quantidade_estoque"]), "unidades em estoque")
    with columns[1]:
        metric_card("Valor líquido Estoque CD", brl(selected["valor_quebra"]), "Valor atual da base")
    with columns[2]:
        metric_card("Valor líquido Ajustes Lojas", brl(selected["valor_identificado"]), "única base de identificação")
    with columns[3]:
        metric_card("Cobertura", percentage(selected["cobertura"]), selected["status"])

    if product_movements.empty:
        st.warning("Nenhuma das movimentações selecionadas foi encontrada para este produto.")
        return

    category_summary = (
        product_movements.groupby("categoria", as_index=False)
        .agg(
            quantidade=("quantidade_movimento_assinada", lambda values: values.sum(min_count=1)),
            valor=("valor_movimento_assinado", "sum"),
            registros=("codigo_produto", "size"),
        )
        .set_index("categoria")
        .reindex(CATEGORIES)
        .reset_index()
    )
    category_summary = category_summary[category_summary["valor"].notna()].copy()
    category_summary["Quantidade"] = category_summary["quantidade"].map(
        lambda value: "—" if pd.isna(value) else integer_br(value)
    )
    category_summary["Valor vinculado"] = category_summary["valor"].map(brl)
    category_summary["% da quebra"] = category_summary["valor"].div(selected["valor_quebra"]).map(percentage)
    category_summary["Participa da identificação"] = category_summary["categoria"].map(
        lambda category: "Sim" if category == "Ajustes Lojas" else "Não"
    )
    display_category = category_summary.rename(columns={"categoria": "Categoria", "registros": "Registros"})[
        ["Categoria", "Quantidade", "Valor vinculado", "% da quebra", "Participa da identificação", "Registros"]
    ]
    st.markdown("#### Movimentações do produto")
    st.caption("Somente Ajustes Lojas participa do cálculo de identificação; as demais movimentações são informativas.")
    st.dataframe(display_category, hide_index=True, width="stretch")

    available_categories = [
        category for category in CATEGORIES if category in set(product_movements["categoria"])
    ]
    tabs = st.tabs(available_categories)
    for tab, category in zip(tabs, available_categories):
        with tab:
            table = detail_table(product_movements, category)
            if category == "Ajustes Lojas":
                st.caption("Detalhamento por loja. Esta movimentação participa da identificação da quebra.")
            elif category == "Inventário":
                st.caption("Detalhamento informativo por CGO. Não participa da identificação da quebra.")
            else:
                st.caption("Detalhamento informativo da movimentação. Não participa da identificação da quebra.")
            st.dataframe(table, hide_index=True, width="stretch", height=min(420, 72 + 36 * len(table)))


def main() -> None:
    st.set_page_config(page_title="Análise de Quebra CD", page_icon="📦", layout="wide")
    inject_css()

    with st.sidebar:
        st.markdown("### Bases de dados")
        st.caption("Envie uma nova versão da planilha ou use a base incluída no projeto.")
        uploaded = st.file_uploader("Planilha de análise", type=["xlsx"], help="Arquivo Excel com as cinco abas esperadas.")
        if uploaded is not None:
            file_bytes = uploaded.getvalue()
            source_label = uploaded.name
        elif DEFAULT_FILE.exists():
            file_bytes = DEFAULT_FILE.read_bytes()
            source_label = DEFAULT_FILE.name
        else:
            st.error("Inclua a planilha em data/Analise CD 989.xlsx ou faça o upload acima.")
            st.stop()

        st.success(f"Base ativa: {source_label}")
        st.divider()
        st.markdown("### Filtros")
        search = st.text_input("Produto ou código", placeholder="Ex.: 2203 ou Itaipava")
        detail_categories = st.multiselect(
            "Movimentações no detalhamento",
            CATEGORIES,
            default=CATEGORIES,
            help=(
                "Controla somente as tabelas do detalhamento. "
                "A identificação continua usando apenas Ajustes Lojas."
            ),
        )
        order_label = st.radio("Ordenar ranking por", ["Maior quantidade", "Maior valor"], horizontal=False)
        top_n = st.slider("Produtos exibidos", 10, 130, 30, step=10)

    try:
        with st.spinner("Lendo e conciliando as abas..."):
            stock, movements_all, quality = prepare_data(file_bytes)
    except Exception as error:
        st.error(f"Não foi possível processar a planilha: {error}")
        st.stop()

    adjustments = movements_all[movements_all["categoria"].eq("Ajustes Lojas")].copy()
    ranking_all = build_ranking(stock, adjustments)

    total_break = ranking_all["valor_quebra"].sum()
    total_identified = adjustments["valor_movimento_assinado"].sum()
    total_gap = ranking_all["saldo_justificar"].sum()
    coverage = total_identified / total_break if total_break else 0

    st.markdown('<div class="eyebrow">Controle de estoque • CD 989</div>', unsafe_allow_html=True)
    st.title("Análise de Quebras - Estoque CD")
    st.markdown(
        '<div class="hero-subtitle">O Dashboard traz a análise dos produtos com maior quebra no "Estoque CD".</div>',
        unsafe_allow_html=True,
    )

    kpi_columns = st.columns(5)
    with kpi_columns[0]:
        metric_card("Valor de quebra total CD", brl(total_break), f"{len(ranking_all)} produtos na Base")
    with kpi_columns[1]:
        metric_card("Total em Ajustes Lojas", brl(total_identified), "valor líquido encontrado nos ajustes")
    with kpi_columns[2]:
        metric_card("Saldo após Ajustes", brl(total_gap), "soma dos saldos por produto")
    with kpi_columns[3]:
        metric_card("Cobertura Ajustes Lojas", percentage(coverage), "Ajustes Lojas ÷ quebra CD")
    with kpi_columns[4]:
        metric_card("Quantidade em estoque", integer_br(ranking_all["quantidade_estoque"].sum()), "unidades na base de referência")

    stores_tab, overview_tab, ranking_tab, quality_tab = st.tabs(
        ["Ajustes Lojas", "Visão geral", "Ranking e justificativas", "Qualidade dos dados"]
    )

    filtered = ranking_all.copy()
    if search.strip():
        token = normalize_label(search)
        mask = filtered.apply(
            lambda row: token in normalize_label(row["produto"]) or token in normalize_label(row["codigo_produto"]), axis=1
        )
        filtered = filtered[mask]
    order_column = "quantidade_estoque" if order_label == "Maior quantidade" else "valor_quebra"
    filtered = filtered.sort_values([order_column, "valor_quebra"], ascending=False).head(top_n).reset_index(drop=True)
    filtered.insert(0, "ranking", range(1, len(filtered) + 1))

    with overview_tab:
        render_informative_movements(stock, movements_all, search, top_n)

    with ranking_tab:
        st.markdown("#### Comparativo Estoque CD × Ajustes Lojas")
        st.caption(
            "O valor identificado considera exclusivamente o valor líquido de Ajustes Lojas. "
            "Clique em uma linha para abrir o detalhamento por loja."
        )
        display = filtered[
            [
                "ranking",
                "codigo_produto",
                "produto",
                "quantidade_estoque",
                "valor_quebra",
                "valor_identificado",
                "cobertura",
            ]
        ].rename(
            columns={
                "ranking": "#",
                "codigo_produto": "Código",
                "produto": "Produto",
                "quantidade_estoque": "Quantidade",
                "valor_quebra": "Valor líquido Estoque CD",
                "valor_identificado": "Valor líquido Ajustes Lojas",
                "cobertura": "Ajustes / Estoque",
            }
        )

        selection = st.dataframe(
            display,
            hide_index=True,
            width="stretch",
            height=min(720, 39 + 35 * max(len(display), 1)),
            on_select="rerun",
            selection_mode="single-row",
            key="ranking_selection",
            column_config={
                "#": st.column_config.NumberColumn(width="small", format="%d"),
                "Código": st.column_config.TextColumn(width="small"),
                "Produto": st.column_config.TextColumn(width="large"),
                "Quantidade": st.column_config.NumberColumn(format="%,.0f"),
                "Valor líquido Estoque CD": st.column_config.NumberColumn(format="R$ %,.2f"),
                "Valor líquido Ajustes Lojas": st.column_config.NumberColumn(format="R$ %,.2f"),
                "Ajustes / Estoque": st.column_config.NumberColumn(format="percent"),
            },
        )

        csv_data = display.to_csv(index=False, sep=";", decimal=",", encoding="utf-8-sig")
        st.download_button(
            "Baixar ranking filtrado (CSV)",
            data=csv_data.encode("utf-8-sig"),
            file_name="comparativo_estoque_cd_ajustes_lojas.csv",
            mime="text/csv",
            width="stretch",
        )

        selected_rows = selection.selection.rows if selection is not None else []
        if selected_rows and not filtered.empty:
            selected = filtered.iloc[selected_rows[0]]
        elif not filtered.empty:
            selected_code = st.selectbox(
                "Ou escolha um produto para detalhar",
                options=filtered["codigo_produto"].tolist(),
                format_func=lambda code: f"{code} · {filtered.loc[filtered['codigo_produto'].eq(code), 'produto'].iloc[0]}",
            )
            selected = filtered[filtered["codigo_produto"].eq(selected_code)].iloc[0]
        else:
            selected = None

        if selected is not None:
            detail_movements = movements_all[movements_all["categoria"].isin(detail_categories)]
            product_movements = detail_movements[
                detail_movements["codigo_produto"].eq(selected["codigo_produto"])
            ]
            render_product_detail(selected, product_movements)

        st.markdown(
            '<p class="method-note"><strong>Regra de identificação:</strong> somente o valor líquido assinado de Ajustes Lojas participa do comparativo, da cobertura e do saldo. Transferências, Devoluções e Inventário são informativos e podem ser ativados ou desativados no filtro de detalhamento.</p>',
            unsafe_allow_html=True,
        )

    with stores_tab:
        render_store_ranking(stock, movements_all, search, top_n)

    with quality_tab:
        st.markdown("#### Cobertura das abas")
        st.caption("Produtos exclusivos das outras abas são contados como ignorados e não entram nos KPIs nem no ranking.")
        st.dataframe(quality, hide_index=True, width="stretch")

        no_evidence = ranking_all[ranking_all["movimentacao_vinculada"].eq(0)][["codigo_produto", "produto", "valor_quebra"]]
        no_evidence = no_evidence.sort_values("valor_quebra", ascending=False)
        no_evidence["valor_quebra"] = no_evidence["valor_quebra"].map(brl)
        st.markdown("#### Produtos do CD sem Ajustes Lojas")
        if no_evidence.empty:
            st.success("Todos os produtos do Estoque CD possuem valores vinculados em Ajustes Lojas.")
        else:
            st.dataframe(
                no_evidence.rename(columns={"codigo_produto": "Código", "produto": "Produto", "valor_quebra": "Valor da quebra"}),
                hide_index=True,
                width="stretch",
            )


if __name__ == "__main__":
    main()
